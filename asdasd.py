import smtplib
import openpyxl
import time


from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

wb = load_workbook(filename = 'flags.xlsx')
ws: Worksheet = wb.active
print(ws['a1'].value)
val = ws['a1'].value

i = 2
while i<18:
    strana = ws.cell(row = i, column=1).value
    adres = ws.cell(row = i, column=4).value

    fromaddr = "artemkartunchikov@mail.ru"
    toaddr = adres
    mypass = "123123"
    print(ws.cell(row=i, column=1).value)
    print(ws.cell(row = i, column=3).value)
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = "The flag of " + strana

    body = "Dear representatives of " + strana + ", my name is Kartunchikov Artem. I study at the Russian Technical University.\nI want to create a discussion club in our university, the purpose of which will be to gain knowledge about the cultural diversity of our planet.\nAnd for this, I dream to get the attributes of all countries. I will be happy if you can send me the flag or coat of arms of your country, even a small one will be great.\nMy address Moskovskaya oblast, Reutov, Lenina 2 32, index 143964\n\n\n\nI wish you all the best\nArtem"

    msg.attach(MIMEText(body, 'plain'))

    server = smtplib.SMTP_SSL('smtp.mail.ru', 465)
    server.login(fromaddr, mypass)
    text = msg.as_string()
    server.sendmail(fromaddr, toaddr, text)
    server.quit()
    time.sleep(2)
    i+=1
